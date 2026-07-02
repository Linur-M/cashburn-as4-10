"""
fx_excel.py — load Bank-of-Israel exchange rates from a USER-PROVIDED Excel sheet and merge
them into the in-memory _FX_LIVE table used by Step 0's FX pipeline.

Expected layout (sheet named like BOI/Rates/שער, else the first sheet):
    Date         USD     EUR     GBP
    2026-03-02   3.70    4.00    4.65
    2026-03-31   3.63    3.92    4.57
    ...
- Date  : ISO "YYYY-MM-DD" or a real Excel date (also accepts DD/MM/YYYY, DD.MM.YYYY).
          Rows with a blank/unparseable date are skipped.
- One column per currency, header = ISO code (USD/EUR/GBP recognised; any 3-letter code kept).
- Values are ILS per 1 unit of that currency (the BOI representative rate). ILS itself is the
  base (=1) and is ignored if present as a column.

When a rates Excel is supplied it is AUTHORITATIVE: it overrides other sources for the dates it
covers. Gaps between supplied dates are handled by the existing get_fx() walkback, and the
month-end anchor (last_boi_trading_day) naturally resolves to the last supplied USD date in the
month.
"""
import re
from datetime import datetime, date

_CCY_RE   = re.compile(r"^[A-Z]{3}$")
_DATE_HDR = re.compile(r"^\s*(date|תאריך)\s*$", re.IGNORECASE)
_DATE_FMTS = ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%d.%m.%Y", "%d-%m-%Y", "%Y/%m/%d")


def _norm_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s:
        return None
    for fmt in _DATE_FMTS:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    m = re.search(r"\d{4}-\d{2}-\d{2}", s)          # last resort: ISO date embedded in text
    return m.group(0) if m else None


def _to_float(v):
    if v is None or v == "":
        return None
    try:
        return float(str(v).replace(",", "").replace("%", ""))
    except ValueError:
        return None


def load_fx_from_excel(path, sheet=None):
    """
    Parse a rates workbook.
    Returns (rates, meta):
      rates = {(YYYY-MM-DD, CCY): float}        — ready to merge into _FX_LIVE
      meta  = {sheet, currencies, n_rows, date_min, date_max, source}
    """
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)

    name = sheet
    if name is None:                                # prefer a rates-looking sheet, else first
        for s in wb.sheetnames:
            if re.search(r"boi|rate|fx|שער", s, re.IGNORECASE):
                name = s
                break
        name = name or wb.sheetnames[0]
    ws = wb[name]
    rows = list(ws.iter_rows(values_only=True))
    fname = path.replace("\\", "/").split("/")[-1]
    if not rows:
        return {}, {"sheet": name, "currencies": [], "n_rows": 0,
                    "date_min": None, "date_max": None,
                    "source": f"BOI_xlsx ({fname} [{name}]) — empty"}

    # locate header row: first row (of the first few) with a Date cell + >=1 currency code
    hdr_i, headers = 0, rows[0]
    for i, r in enumerate(rows[:5]):
        cells = [("" if c is None else str(c).strip()) for c in r]
        if any(_DATE_HDR.match(c) for c in cells) and \
           any(_CCY_RE.match(c.upper()) for c in cells):
            hdr_i, headers = i, r
            break

    date_col, cols = None, {}
    for j, h in enumerate(headers):
        hs = "" if h is None else str(h).strip()
        if _DATE_HDR.match(hs):
            date_col = j
        elif _CCY_RE.match(hs.upper()) and hs.upper() != "ILS":
            cols[hs.upper()] = j

    rates, dmin, dmax, n = {}, None, None, 0
    for r in rows[hdr_i + 1:]:
        if date_col is None or date_col >= len(r):
            continue
        d = _norm_date(r[date_col])
        if not d:
            continue
        wrote = False
        for ccy, j in cols.items():
            if j < len(r):
                val = _to_float(r[j])
                if val and val > 0:
                    rates[(d, ccy)] = val
                    wrote = True
        if wrote:
            n += 1
            dmin = d if (dmin is None or d < dmin) else dmin
            dmax = d if (dmax is None or d > dmax) else dmax

    meta = {"sheet": name, "currencies": sorted(cols.keys()), "n_rows": n,
            "date_min": dmin, "date_max": dmax,
            "source": f"BOI_xlsx ({fname} [{name}])"}
    return rates, meta


def merge_into_fx_live(fx_live, rates, override=True):
    """Merge Excel rates into _FX_LIVE. override=True → the uploaded file wins on conflicts."""
    for k, v in rates.items():
        if override or k not in fx_live:
            fx_live[k] = v
    return fx_live


def month_coverage(rates_or_live, month_str, ccy="USD"):
    """Sorted days of month_str (YYYY-MM) that carry a `ccy` quote — for the pre-flight line."""
    return sorted(d for (d, c) in rates_or_live if c == ccy and d[:7] == month_str)


def find_rates_excel(upload_paths):
    """Pick the most likely rates workbook from a list of uploaded paths (by filename)."""
    for p in upload_paths:
        base = p.replace("\\", "/").split("/")[-1].lower()
        if base.endswith((".xlsx", ".xlsm")) and re.search(r"boi|rate|fx|שער|מטבע", base):
            return p
    return None
